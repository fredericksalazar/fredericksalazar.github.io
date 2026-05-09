"""Ingester DANE — Mercado laboral (GEIH) mensual.

DANE publica el boletín mensual de la Gran Encuesta Integrada de Hogares (GEIH)
con anexos Excel que contienen las series históricas nacionales de:
  - Tasa global de participación (TGP)
  - Tasa de ocupación (TO)
  - Tasa de desempleo (TD)
  - Proporción de informalidad (rama "informal-formalidad")

Patrón de URL (anexo principal mensual nacional):
  https://www.dane.gov.co/files/operaciones/GEIH/anex-GEIH-{mes}{anio}.xlsx

Fuente histórica complementaria (one-time):
  https://www.dane.gov.co/files/investigaciones/boletines/ech/nuevo-enfoque-conceptual-metodologico-2018/anexo-mercado-laboral-segun-proyecciones-CNPV2018.xlsx
  (serie completa 2001-2021 con proyecciones CNPV 2018)

Donde {mes} es la abreviatura en minúscula de tres letras (ene, feb, ..., dic)
y {anio} es el año a 4 dígitos. El archivo del mes M sale entre los días 28-31
del mes M+1.

Notas operativas:
- DANE puede mover/renombrar archivos: el ingester degrada elegantemente
  retrocediendo hasta 4 meses si la URL del mes objetivo no responde.
- Hojas comunes: 'tasas nacional', 'TGP_TO_TD', 'Anexo nacional'. La extracción
  exacta puede requerir ajustes si DANE cambia el layout.
- Los datos de informalidad se publican en un anexo separado:
  anex-GEIH-formalidad-{mes}{anio}.xlsx (si existe).

Este módulo expone fetch() pero retorna un DataFrame vacío con columnas
correctas si la fuente no está disponible: el resto del pipeline puede
generar el JSON usando solo el histórico de World Bank.
"""

from __future__ import annotations

import io
import logging
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import requests

BASE_URL = "https://www.dane.gov.co/files/operaciones/GEIH"
TIMEOUT_SECONDS = 30
HEADERS = {"User-Agent": "observatorio-economico-colombia/0.1"}

FUENTE = "dane"
INDICADOR = "empleo"
FUENTE_NOMBRE = "DANE — Departamento Administrativo Nacional de Estadística"
FUENTE_URL = (
    "https://www.dane.gov.co/index.php/estadisticas-por-tema/"
    "mercado-laboral/empleo-y-desempleo"
)
INDICADOR_NOMBRE = "Mercado laboral — Gran Encuesta Integrada de Hogares (GEIH)"

MESES_ABREV = [
    "ene", "feb", "mar", "abr", "may", "jun",
    "jul", "ago", "sep", "oct", "nov", "dic",
]

COLUMNAS = ["periodo", "tasa_desempleo", "tgp", "to", "subempleo", "informalidad"]

logger = logging.getLogger("observatorio.dane_empleo")


def _build_url(year: int, month: int) -> str:
    abrev = MESES_ABREV[month - 1]
    return f"{BASE_URL}/anex-GEIH-{abrev}{year}.xlsx"


HISTORICO_URL = (
    "https://www.dane.gov.co/files/investigaciones/boletines/ech/"
    "nuevo-enfoque-conceptual-metodologico-2018/"
    "anexo-mercado-laboral-segun-proyecciones-CNPV2018.xlsx"
)


def _download_historico() -> bytes | None:
    """Descarga el archivo histórico CNPV 2018 (serie 2001-2021)."""
    return _download(HISTORICO_URL)


def _download(url: str) -> bytes | None:
    """Descarga URL. Usa curl como método principal (más robusto SSL en macOS)."""
    import subprocess
    try:
        result = subprocess.run(
            ["curl", "-sS", "-L", "--max-time", str(TIMEOUT_SECONDS),
             "-w", "\\n__HTTP_CODE__:%{http_code}",
             "-H", f"User-Agent: {HEADERS['User-Agent']}", url],
            capture_output=True, timeout=TIMEOUT_SECONDS + 5,
        )
        if result.returncode == 0 and result.stdout:
            # Extraer HTTP status code del output (añadido por -w)
            stdout = result.stdout
            http_marker = b"__HTTP_CODE__:"
            if http_marker in stdout:
                marker_pos = stdout.rfind(http_marker)
                http_code_str = stdout[marker_pos + len(http_marker):].strip()
                stdout = stdout[:marker_pos]
                try:
                    if int(http_code_str) >= 400:
                        logger.debug("DANE GEIH HTTP %s: %s", http_code_str, url)
                        return None
                except ValueError:
                    pass
            if stdout:
                return stdout
    except Exception:
        pass

    # Fallback a requests
    try:
        resp = requests.get(url, timeout=TIMEOUT_SECONDS, headers=HEADERS)
    except requests.RequestException as e:
        logger.debug("DANE GEIH no responde (%s): %s", url, e)
        return None
    if resp.status_code == 200 and resp.content:
        return resp.content
    return None


def _find_latest_available(today: date | None = None) -> tuple[int, int, bytes] | None:
    """Busca el anexo GEIH más reciente disponible, hasta 4 meses atrás."""
    if today is None:
        today = date.today()
    cursor = date(today.year, today.month, 1)
    for _ in range(4):
        if cursor.month == 1:
            cursor = date(cursor.year - 1, 12, 1)
        else:
            cursor = date(cursor.year, cursor.month - 1, 1)
        content = _download(_build_url(cursor.year, cursor.month))
        if content:
            return cursor.year, cursor.month, content
    return None


def _empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=COLUMNAS)


def _parse_anexo(content: bytes) -> pd.DataFrame:
    """Extrae la serie nacional del anexo GEIH (formato wide: meses como columnas).

    El anexo usa la hoja 'Total nacional' con estructura:
      Row 10: título
      Row 11: 'Concepto' | 2001 | None | ... | 2002 | ... (años, merged)
      Row 12: None | 'Ene' | 'Feb' | ... | 'Dic' | 'Ene' | ...
      Row 13+:  TGP / TO / TD / etc.

    Si el layout cambió, retorna DataFrame vacío.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        logger.warning("No se pudo abrir anexo GEIH: %s", e)
        return _empty_df()

    # Buscar hoja de datos (puede tener distintos nombres entre anexos)
    target_sheet = None
    for candidate in ("Total nacional", "Tnal mensual", "Total nacional mensual"):
        if candidate in wb.sheetnames:
            target_sheet = candidate
            break
    if target_sheet is None:
        logger.warning("Hoja de datos nacional no encontrada en anexo GEIH. Hojas: %s",
                       wb.sheetnames[:6])
        return _empty_df()

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 15:
        return _empty_df()

    # Encontrar fila de años (row 11 0-indexed, pero puede variar)
    year_row_idx = None
    month_row_idx = None
    for i in range(10, min(20, len(rows))):
        row = rows[i]
        if row and row[0] and str(row[0]).strip() == "Concepto":
            year_row_idx = i
            month_row_idx = i + 1
            break

    if year_row_idx is None:
        logger.warning("Fila 'Concepto' no encontrada en anexo GEIH")
        return _empty_df()

    year_row = rows[year_row_idx]
    month_row = rows[month_row_idx]

    # Construir periodos (YYYY-MM) para cada columna de datos
    periodos: list[str] = []
    current_year: int | None = None
    for col in range(1, len(year_row)):
        year_cell = year_row[col]
        month_cell = month_row[col] if col < len(month_row) else None
        if year_cell is not None:
            try:
                current_year = int(year_cell)
            except (ValueError, TypeError):
                pass
        if current_year is None or month_cell is None:
            continue
        month_str = str(month_cell).strip().lower()
        if month_str not in _MESES_ABREV_MAP:
            continue
        month = _MESES_ABREV_MAP[month_str]
        periodos.append(f"{current_year:04d}-{month:02d}")

    if not periodos:
        logger.warning("No se pudieron extraer periodos del anexo GEIH")
        return _empty_df()

    # Buscar filas de conceptos (TGP, TO, TD, TS)
    concept_map: dict[str, int] = {}  # key → index en filas
    for i in range(year_row_idx + 1, min(year_row_idx + 50, len(rows))):
        row = rows[i]
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().upper()
        # Detener al llegar a "Total Cabeceras" (siguiente sección)
        if "TOTAL CABECERAS" in label:
            break
        if "TASA GLOBAL DE PARTICIPACIÓN" in label or label == "TGP":
            concept_map["tgp"] = i
        elif "TASA DE OCUPACIÓN" in label or label == "TO":
            concept_map["to"] = i
        elif "TASA DE DESOCUPACIÓN" in label or "DESEMPLEO" in label or label == "TD":
            concept_map["tasa_desempleo"] = i
        elif "TASA DE SUBOCUPACIÓN" in label or label == "TS":
            concept_map["subempleo"] = i

    if len(concept_map) < 2:
        logger.warning("No se encontraron suficientes conceptos (TGP/TO/TD) en anexo GEIH")
        return _empty_df()

    # Extraer valores: para cada periodo, leer el valor en la columna correspondiente
    records: list[dict] = []
    for col_idx, periodo in enumerate(periodos):
        excel_col = col_idx + 1  # columna 0 es concepto
        rec: dict = {"periodo": periodo, "informalidad": None}
        for key, row_idx in concept_map.items():
            cell = rows[row_idx][excel_col] if excel_col < len(rows[row_idx]) else None
            rec[key] = round(float(cell), 2) if isinstance(cell, (int, float)) else None
        if any(rec.get(k) is not None for k in ("tgp", "to", "tasa_desempleo")):
            records.append(rec)

    if not records:
        return _empty_df()

    return (
        pd.DataFrame(records)
        .sort_values("periodo")
        .reset_index(drop=True)
    )


# Mapa de abreviaturas de meses usadas por DANE
_MESES_ABREV_MAP = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def fetch() -> pd.DataFrame:
    """Descarga el anexo GEIH más reciente y devuelve la serie nacional.

    Estrategia:
    1. Descarga el histórico CNPV 2018 (serie 2001-2021) como base.
    2. Descarga el último anexo mensual para extender la serie al presente.
    3. Merge: los datos mensuales más recientes prevalecen sobre los históricos.

    Si la fuente no está disponible o el layout cambió, retorna DataFrame vacío
    con las columnas esperadas — el pipeline degrada a usar solo World Bank.
    """
    dfs: list[pd.DataFrame] = []

    # 1. Histórico CNPV 2018 (2001-2021)
    hist_content = _download_historico()
    if hist_content:
        df_hist = _parse_anexo(hist_content)
        if not df_hist.empty:
            dfs.append(df_hist)
            logger.info("DANE GEIH histórico CNPV 2018: %d filas", len(df_hist))
    else:
        logger.info("DANE GEIH histórico CNPV 2018 no disponible")

    # 2. Último anexo mensual
    found = _find_latest_available()
    if found:
        year, month, content = found
        df_mensual = _parse_anexo(content)
        if not df_mensual.empty:
            df_mensual.attrs["last_published"] = f"{year:04d}-{month:02d}"
            dfs.append(df_mensual)
            logger.info("DANE GEIH anexo %d-%02d: %d filas", year, month, len(df_mensual))
    else:
        logger.info("DANE GEIH anexo mensual no disponible")

    if not dfs:
        logger.info("DANE GEIH no disponible — degradando a histórico WB")
        return _empty_df()

    # 3. Merge: último anexo prevalece sobre histórico en periodos solapados
    if len(dfs) == 1:
        df = dfs[0]
    else:
        df = pd.concat(dfs, ignore_index=True)
        df = df.drop_duplicates(subset=["periodo"], keep="last")

    df = df.sort_values("periodo").reset_index(drop=True)
    return df


def save_snapshot(df: pd.DataFrame, raw_root: Path) -> Path | None:
    if df.empty:
        return None
    out_dir = raw_root / FUENTE / INDICADOR
    out_dir.mkdir(parents=True, exist_ok=True)
    ultimo = df["periodo"].iloc[-1]
    path = out_dir / f"{INDICADOR}_{ultimo}.csv"
    df.to_csv(path, index=False)
    return path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    df = fetch()
    if df.empty:
        print("DANE GEIH: sin datos (fuente no disponible o layout cambiado).")
    else:
        print(f"Filas: {len(df)} — último: {df['periodo'].iloc[-1]}")
        print(df.tail(6))
