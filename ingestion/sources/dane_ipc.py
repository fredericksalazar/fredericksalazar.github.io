"""Ingester DANE — IPC mensual y anual (inflación).

DANE no expone API. Publica anexos Excel mensuales con la serie histórica completa.
Patrón de URL:
  https://www.dane.gov.co/files/operaciones/IPC/{mes}{anio}/anex-IPC-{tipo}-{mes}{anio}.xlsx

Cada archivo mensual contiene la serie completa desde 2003-01.
"""

from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import requests

from ingestion.exceptions import FuenteNoDisponibleError

BASE_URL = "https://www.dane.gov.co/files/operaciones/IPC"
TIMEOUT_SECONDS = 30
HEADERS = {"User-Agent": "observatorio-economico-colombia/0.1"}

FUENTE = "dane"
INDICADOR = "inflacion"
FUENTE_NOMBRE = "DANE — Departamento Administrativo Nacional de Estadística"
FUENTE_URL = "https://www.dane.gov.co/index.php/estadisticas-por-tema/precios-y-costos/indice-de-precios-al-consumidor-ipc"
INDICADOR_NOMBRE = "Índice de Precios al Consumidor (IPC)"

MESES_ABREV = [
    "ene", "feb", "mar", "abr", "may", "jun",
    "jul", "ago", "sep", "oct", "nov", "dic",
]
MESES_NOMBRE = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
    "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
    "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12,
}


def _build_url(tipo: str, year: int, month: int) -> str:
    """tipo ∈ {'Variacion', 'Indices'}."""
    abrev = MESES_ABREV[month - 1]
    folder = f"{abrev}{year}"
    fname = f"anex-IPC-{tipo}-{abrev}{year}.xlsx"
    return f"{BASE_URL}/{folder}/{fname}"


def _download_xlsx(url: str) -> bytes | None:
    try:
        resp = requests.get(url, timeout=TIMEOUT_SECONDS, headers=HEADERS)
    except requests.RequestException:
        return None
    if resp.status_code == 200 and resp.content:
        return resp.content
    return None


def _find_latest_available(today: date | None = None) -> tuple[int, int, bytes, bytes]:
    """Busca el archivo DANE más reciente disponible, retrocediendo desde today.

    DANE publica entre días 5-9 del mes siguiente. El archivo del mes M
    aparece en folder/nombre con el mes M (ej. mar2026 = datos hasta marzo).
    """
    if today is None:
        today = date.today()

    # Intentar mes pasado, mes anteanterior, etc., hasta 3 meses atrás
    cursor = date(today.year, today.month, 1)
    for _ in range(4):
        # Retroceder un mes
        if cursor.month == 1:
            cursor = date(cursor.year - 1, 12, 1)
        else:
            cursor = date(cursor.year, cursor.month - 1, 1)

        var_bytes = _download_xlsx(_build_url("Variacion", cursor.year, cursor.month))
        idx_bytes = _download_xlsx(_build_url("Indices", cursor.year, cursor.month))
        if var_bytes and idx_bytes:
            return cursor.year, cursor.month, var_bytes, idx_bytes

    raise FuenteNoDisponibleError(
        "DANE",
        "No se encontró archivo IPC en los últimos 4 meses",
    )


def _despivot(content: bytes, sheet_name: str, value_col: str) -> pd.DataFrame:
    """Convierte la matriz mes×año del DANE a serie temporal."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise FuenteNoDisponibleError("DANE", f"Hoja '{sheet_name}' no encontrada")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (i for i, r in enumerate(rows) if r and r[0] == "Mes"),
        None,
    )
    if header_idx is None:
        raise FuenteNoDisponibleError("DANE", f"Header 'Mes' no encontrado en {sheet_name}")

    header = rows[header_idx]
    years = [int(y) for y in header[1:] if isinstance(y, int)]

    records: list[dict] = []
    for row in rows[header_idx + 1:]:
        if not row or not isinstance(row[0], str):
            continue
        mes_nombre = row[0].strip()
        if mes_nombre not in MESES_NOMBRE:
            continue
        mes_num = MESES_NOMBRE[mes_nombre]
        for col_idx, year in enumerate(years, start=1):
            valor = row[col_idx] if col_idx < len(row) else None
            if valor is None or not isinstance(valor, (int, float)):
                continue
            periodo = f"{year:04d}-{mes_num:02d}"
            records.append({"periodo": periodo, value_col: float(valor)})

    if not records:
        raise FuenteNoDisponibleError("DANE", f"Sin datos en {sheet_name}")

    return (
        pd.DataFrame(records)
        .drop_duplicates(subset=["periodo"])
        .sort_values("periodo")
        .reset_index(drop=True)
    )


def fetch(start_year: int = 2003) -> pd.DataFrame:
    """Descarga la serie histórica del IPC.

    Returns:
        DataFrame con columnas [periodo, inflacion_mensual, inflacion_anual]
        donde inflacion_anual = (idx_t / idx_t-12 - 1) * 100.
    """
    year, month, var_bytes, idx_bytes = _find_latest_available()

    df_var = _despivot(var_bytes, "VarNal", "inflacion_mensual")
    df_idx = _despivot(idx_bytes, "IndicesIPC", "indice")

    df_idx["indice_lag12"] = df_idx["indice"].shift(12)
    df_idx["inflacion_anual"] = (
        (df_idx["indice"] / df_idx["indice_lag12"] - 1) * 100
    ).round(2)

    df = df_var.merge(
        df_idx[["periodo", "inflacion_anual"]],
        on="periodo",
        how="left",
    )

    df = df[df["periodo"] >= f"{start_year}-01"].reset_index(drop=True)
    df.attrs["last_published"] = f"{year:04d}-{month:02d}"
    return df


def save_snapshot(df: pd.DataFrame, raw_root: Path) -> Path:
    """Guarda snapshot crudo en data/raw/{fuente}/{indicador}/{indicador}_{yyyy-mm}.csv."""
    out_dir = raw_root / FUENTE / INDICADOR
    out_dir.mkdir(parents=True, exist_ok=True)
    ultimo = df["periodo"].iloc[-1]
    path = out_dir / f"{INDICADOR}_{ultimo}.csv"
    df.to_csv(path, index=False)
    return path


if __name__ == "__main__":
    df = fetch(start_year=2003)
    print(f"Último archivo DANE publicado: {df.attrs.get('last_published')}")
    print(f"Filas: {len(df)}")
    print(df.head())
    print("...")
    print(df.tail(10))
