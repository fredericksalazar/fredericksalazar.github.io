"""Ingester DANE — Comercio Exterior (Balanza Comercial + OMC + Tradicionales).

Fuentes:
  - Exportaciones OMC: anex-EXPORTACIONES-SerieTotalesProductosOMCAgrCUCIR3-{mes}{año}.xlsx
  - Importaciones OMC: anex-IMP-ProductosOMCAgrCUCI3-{mes}{año}.xlsx
  - Café/carbón/petróleo: anex-EXPORTACIONES-SerieCafeCarbonPetroleoNotradicionales-{mes}{año}.xlsx

Degrada gracefulmente si los archivos no están disponibles.
"""

from __future__ import annotations

import io
import subprocess
import logging
from datetime import date, datetime

import openpyxl
import pandas as pd

BASE_URLS = {
    "expo": "https://www.dane.gov.co/files/operaciones/EXPORTACIONES",
    "imp": "https://www.dane.gov.co/files/operaciones/IMP",
}

MESES_ABREV = [
    "ene", "feb", "mar", "abr", "may", "jun",
    "jul", "ago", "sep", "oct", "nov", "dic",
]

TIMEOUT = 30
HEADERS = {"User-Agent": "observatorio-economico-colombia/0.1"}

logger = logging.getLogger("observatorio.dane_comercio")


def _download(url: str) -> bytes | None:
    try:
        result = subprocess.run(
            ["curl", "-sS", "-L", "--max-time", str(TIMEOUT),
             "-w", "\\n__HTTP__:%{http_code}",
             "-H", f"User-Agent: {HEADERS['User-Agent']}", url],
            capture_output=True, timeout=TIMEOUT + 5,
        )
        if result.returncode == 0 and result.stdout:
            stdout = result.stdout
            marker = b"__HTTP__:"
            if marker in stdout:
                pos = stdout.rfind(marker)
                code = stdout[pos + len(marker):].strip()
                stdout = stdout[:pos]
                if int(code) >= 400:
                    return None
            if stdout:
                return stdout
    except Exception:
        pass
    return None


def _find_latest_anexo(base: str, pattern: str, today: date | None = None) -> bytes | None:
    if today is None:
        today = date.today()
    cursor = date(today.year, today.month, 1)
    for _ in range(4):
        if cursor.month == 1:
            cursor = date(cursor.year - 1, 12, 1)
        else:
            cursor = date(cursor.year, cursor.month - 1, 1)
        abrev = MESES_ABREV[cursor.month - 1]
        url = f"{base}/{pattern.format(mes=abrev, anio=cursor.year)}"
        content = _download(url)
        if content:
            return content
    return None


def _parse_omc_excel(content: bytes) -> list[dict]:
    """Parsea Excel OMC (export o import) con estructura:
    Row 10/11: header con 'Total' + 4 grupos
    Rows siguientes: datetime + valores en miles USD FOB/CIF
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return []

    if "Grupos OMC" not in wb.sheetnames:
        return []

    ws = wb["Grupos OMC"]
    rows = list(ws.iter_rows(values_only=True))

    # Encontrar fila header (contiene 'Total' o 'Agropecuarios')
    header_idx = None
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        vals = [str(c).strip() if c else "" for c in row]
        if any("Total" in v or "Agropecuario" in v for v in vals):
            header_idx = i
            break

    if header_idx is None:
        return []

    records = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        # El periodo es datetime en col 0
        periodo = row[0]
        if isinstance(periodo, (datetime,)):
            periodo_str = f"{periodo.year:04d}-{periodo.month:02d}"
        else:
            continue

        # Columnas: 0=periodo, 1=total, 2=agro, 3=comb, 4=manu, 5=otros
        rec = {"periodo": periodo_str}
        grupos = ["total", "agropecuarios", "combustibles", "manufacturas", "otros"]
        for j, key in enumerate(grupos):
            v = row[j + 1] if j + 1 < len(row) else None
            rec[key] = round(float(v), 2) if isinstance(v, (int, float)) else None
        records.append(rec)

    return records


def _parse_tradicionales_excel(content: bytes) -> list[dict]:
    """Parsea Excel de productos tradicionales:
    Row 13+: datetime | None | cafe_usd | cafe_tons | ... | carbon_usd | ...
    Columnas: A=mes, C/cafe_usd, F=carbon_usd, I=petroleo_usd, L=ferroniquel_usd
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return []

    if "Tra y Notra" not in wb.sheetnames:
        return []

    ws = wb["Tra y Notra"]
    rows = list(ws.iter_rows(values_only=True))

    # Datos desde row 13 (0-indexed) hasta el final
    records = []
    for row in rows[13:]:
        if not row or row[0] is None:
            continue
        periodo = row[0]
        if isinstance(periodo, (datetime,)):
            periodo_str = f"{periodo.year:04d}-{periodo.month:02d}"
        else:
            continue

        col_map = {
            "cafe": 2,           # C
            "carbon": 5,         # F
            "petroleo": 8,       # I
            "ferroniquel": 11,   # L
            "no_tradicionales": None,  # Calculado: total - tradicionales
            "total_tradicionales": 14,  # O
        }

        rec = {"periodo": periodo_str}
        for key, col in col_map.items():
            if col is not None and col < len(row):
                v = row[col]
                rec[key] = round(float(v), 2) if isinstance(v, (int, float)) else None
            else:
                rec[key] = None
        records.append(rec)

    return records


def fetch_omc_exportaciones() -> list[dict]:
    """Descarga y parsea desglose OMC de exportaciones."""
    content = _find_latest_anexo(
        BASE_URLS["expo"],
        "anex-EXPORTACIONES-SerieTotalesProductosOMCAgrCUCIR3-{mes}{anio}.xlsx",
    )
    if not content:
        return []
    return _parse_omc_excel(content)


def fetch_omc_importaciones() -> list[dict]:
    """Descarga y parsea desglose OMC de importaciones."""
    content = _find_latest_anexo(
        BASE_URLS["imp"],
        "anex-IMP-ProductosOMCAgrCUCI3-{mes}{anio}.xlsx",
    )
    if not content:
        return []
    return _parse_omc_excel(content)


def fetch_productos_tradicionales() -> list[dict]:
    """Descarga y parsea café, carbón, petróleo, ferroníquel."""
    content = _find_latest_anexo(
        BASE_URLS["expo"],
        "anex-EXPORTACIONES-SerieCafeCarbonPetroleoNotradicionales-{mes}{anio}.xlsx",
    )
    if not content:
        return []
    return _parse_tradicionales_excel(content)


def resumir_por_anio(records: list[dict]) -> dict:
    """Convierte serie mensual a resumen anual: agrupa por año y suma valores."""
    if not records:
        return {}

    df = pd.DataFrame(records)
    df["year"] = df["periodo"].str[:4]
    value_cols = [c for c in ["total", "agropecuarios", "combustibles", "manufacturas", "otros", "cafe", "carbon", "petroleo", "ferroniquel", "no_tradicionales", "total_tradicionales"] if c in df.columns]

    result: dict = {}
    for year, group in df.groupby("year"):
        year_data: dict = {}
        for col in value_cols:
            vals = group[col].dropna()
            if len(vals) > 0:
                year_data[col] = round(vals.sum(), 2)
            else:
                year_data[col] = None
        result[str(year)] = year_data

    return result
